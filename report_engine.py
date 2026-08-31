"""
report_engine.py
-----------------
Core logic for converting a raw "Assessment Capture Report" export into the
NGO x Module workflow summary format (as used in SAT_Work_Flow_Data_Monitoring_Sheet).

No UI code here — app.py (Streamlit) imports these functions.
"""

from __future__ import annotations

import io
import json
import re
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# 1. Column normalisation
# --------------------------------------------------------------------------
# Raw exports have slightly different header text/whitespace across pulls
# (e.g. "Parameter" vs "Question", trailing spaces). We map every known
# variant to one canonical internal name.
COLUMN_ALIASES = {
    "group": "Group",
    "ngo name": "NGO Name",
    "submission status": "Submission Status",
    "group submission status": "Group Submission Status",
    "date": "Date",
    "category": "Category",
    "parameter": "Question",
    "question": "Question",
    "answer": "Answer",
    "answer value": "Answer Value",
    "document description": "Document Description",
    "file name": "File Name",
    "rejection comments": "Rejection Comments",
    "feedback": "Feedback",
}

# Columns whose values only appear on the first row of a block in the raw
# export (merged cells) and must be forward-filled.
FORWARD_FILL_COLUMNS = [
    "Group",
    "NGO Name",
    "Submission Status",
    "Group Submission Status",
    "Date",
    "Category",
]

REQUIRED_COLUMNS = [
    "NGO Name",
    "Category",
    "Question",
    "Answer",
    "File Name",
    "Feedback",
]

# Text columns that get whitespace-stripped after loading, so values like
# "Governance " or "Priscilla Centre " match cleanly.
STRIP_COLUMNS = ["NGO Name", "Category"]


def _normalise_header(raw_header: str) -> str:
    key = str(raw_header).strip().lower()
    key = re.sub(r"\s+", " ", key)
    return COLUMN_ALIASES.get(key, str(raw_header).strip())


def load_raw(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Load a raw Assessment Capture Report (.xls or .xlsx) into a clean,
    forward-filled DataFrame with canonical column names.
    """
    ext = Path(filename).suffix.lower()
    buf = io.BytesIO(file_bytes)

    if ext == ".xls":
        # Legacy BIFF format -> needs xlrd (pip install xlrd)
        df = pd.read_excel(buf, engine="xlrd", header=None)
    else:
        df = pd.read_excel(buf, engine="openpyxl", header=None)

    # Find the header row: the first row that contains "NGO Name" (or its
    # normalised equivalent) — the exports have 1-3 title rows above it.
    header_row_idx = None
    for i in range(min(10, len(df))):
        row_vals = [str(v).strip().lower() for v in df.iloc[i].tolist()]
        if any("ngo name" in v for v in row_vals):
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError(
            "Could not find the header row (looking for an 'NGO Name' column) "
            "in the first 10 rows of the uploaded file."
        )

    headers = [_normalise_header(h) for h in df.iloc[header_row_idx].tolist()]
    data = df.iloc[header_row_idx + 1 :].copy()
    data.columns = headers

    # Drop fully-empty rows
    data = data.dropna(how="all")

    # Dedupe columns (in case a header name repeats)
    data = data.loc[:, ~data.columns.duplicated()]

    missing = [c for c in REQUIRED_COLUMNS if c not in data.columns]
    if missing:
        raise ValueError(f"Raw file is missing expected column(s): {missing}")

    # Forward-fill the merged-cell columns
    for col in FORWARD_FILL_COLUMNS:
        if col in data.columns:
            data[col] = data[col].ffill()

    # Strip stray leading/trailing whitespace from key text values. Raw
    # exports sometimes have e.g. "Governance " or "Priscilla Centre " with
    # a trailing space. Left unstripped, these silently mismatch against
    # values a user types by hand into the Program Config (category_order,
    # NGO-specific settings, etc.), causing that data to disappear from the
    # final report with no error.
    for col in STRIP_COLUMNS:
        if col in data.columns:
            data[col] = data[col].apply(lambda v: v.strip() if isinstance(v, str) else v)

    data = data.reset_index(drop=True)
    return data


# --------------------------------------------------------------------------
# 2. Program configuration (Program / IM / Trainer names — the part that
#    varies across programs & states)
# --------------------------------------------------------------------------
@dataclass
class ProgramConfig:
    program_label: str = "Saksham"                 # e.g. "Saksham", "Niranthara"
    state_code: str = "KA"                          # e.g. KA, TN, JH, OD, NNE
    month_label: str = "Jul"                        # e.g. Jul, Aug
    assessment_name: str = "Assessment Name: Saksham B#1"
    report_date: str = ""                           # e.g. "6th July, 2026"
    im_names: list[str] = field(default_factory=list)       # Implementation Managers
    trainer_names: list[str] = field(default_factory=list)  # Trainers
    # Order in which module/category columns should appear in the final
    # sheet. Leave empty to auto-detect from the raw file (first-seen order).
    category_order: list[str] = field(default_factory=list)
    # Total number of parameters per category, used for % calculations.
    # Leave empty to auto-count distinct Questions per Category from the
    # raw file itself.
    parameters_per_category: dict[str, int] = field(default_factory=dict)

    @property
    def sheet_name(self) -> str:
        name = f"{self.state_code}_{self.program_label}_{self.month_label}"
        return name[:31]  # Excel sheet name limit

    def to_json(self) -> str:
        return json.dumps(asdict(self), indent=2)

    @staticmethod
    def from_json(text: str) -> "ProgramConfig":
        return ProgramConfig(**json.loads(text))


PROGRAMS_DIR = Path(__file__).parent / "programs"
PROGRAMS_DIR.mkdir(exist_ok=True)


def list_saved_programs() -> list[str]:
    return sorted(p.stem for p in PROGRAMS_DIR.glob("*.json"))


def save_program(config: ProgramConfig) -> Path:
    path = PROGRAMS_DIR / f"{config.sheet_name}.json"
    path.write_text(config.to_json())
    return path


def load_program(name: str) -> ProgramConfig:
    path = PROGRAMS_DIR / f"{name}.json"
    return ProgramConfig.from_json(path.read_text())


# --------------------------------------------------------------------------
# 3. Reviewer extraction + default accept/reject classification
# --------------------------------------------------------------------------
NAME_LINE_RE = re.compile(r"^\s*([A-Za-z][A-Za-z .]{1,40}?)\s*:\s*(.*)$")

# Words/phrases that signal a reviewer is accepting the evidence. Matched as
# whole words, case-insensitive. "this is ok/okay/accepted/approved" reads
# naturally against these same word-level patterns.
_ACCEPT_WORDS = r"(?:ok|okay|accepted|approved|fine|good|correct|verified|valid)"
_ACCEPT_RE = re.compile(rf"\b{_ACCEPT_WORDS}\b", re.IGNORECASE)
# Catches "not okay", "not accepted", etc. so a negated accept-word isn't
# mistaken for acceptance.
_NEGATED_ACCEPT_RE = re.compile(
    rf"\bnot\s+(?:really\s+|quite\s+|fully\s+)?{_ACCEPT_WORDS}\b", re.IGNORECASE
)


def _is_acceptance_comment(text: str) -> bool:
    """True if reviewer comment text reads as an acceptance (e.g. 'Okay',
    'This is accepted', 'approved') rather than a rejection/correction note.
    """
    if not text:
        return False
    if _NEGATED_ACCEPT_RE.search(text):
        return False
    return bool(_ACCEPT_RE.search(text))


def extract_reviewer_comments(feedback: Optional[str]) -> list[tuple[str, str]]:
    """Parse a Feedback cell into a list of (name, comment) pairs based on
    the "Name: comment" convention used in the exports (one per line).
    """
    if not feedback or not isinstance(feedback, str):
        return []
    pairs = []
    for line in feedback.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = NAME_LINE_RE.match(line)
        if m:
            pairs.append((m.group(1).strip(), m.group(2).strip()))
    return pairs


def classify_row(answer: Optional[str], feedback: Optional[str], config: ProgramConfig):
    """Return a dict with default reviewer-role flags and accept/reject guess
    for one raw row. This is a DEFAULT/best-guess only — the app surfaces
    these as editable columns because real accept/reject calls require
    reading the evidence document and cannot be inferred from text alone.

    Reviewed = at least one comment came from someone in config.im_names or
    config.trainer_names. A comment from anyone NOT on either list does not
    count as reviewed — the row stays Accepted=0, Rejected=0 even if a
    comment is present.
    Accepted = the IM/Trainer comment(s) contain acceptance language (okay,
    accepted, approved, ...). Reviewed-but-not-accepted = Rejected.

    `answer` (the NGO's own Yes/No self-report) is accepted for call-site
    compatibility but is no longer used to guess Accepted/Rejected.
    """
    comments = extract_reviewer_comments(feedback)

    im_set = {n.strip().lower() for n in config.im_names}
    trainer_set = {n.strip().lower() for n in config.trainer_names}

    reviewer_names = [name for name, _ in comments]
    reviewed_by_im = any(name.strip().lower() in im_set for name in reviewer_names)
    reviewed_by_trainer = any(name.strip().lower() in trainer_set for name in reviewer_names)

    # Only comments from a configured IM or Trainer count toward the
    # reviewed/accepted/rejected guess. A comment from someone not on either
    # list (e.g. a typo'd name, or someone outside the program) is ignored
    # for this purpose.
    official_comments = [
        (name, comment)
        for name, comment in comments
        if name.strip().lower() in im_set or name.strip().lower() in trainer_set
    ]
    has_feedback = len(official_comments) > 0

    if not has_feedback:
        # No comment from a configured IM/Trainer -> not yet reviewed, so no
        # guess is made.
        accepted, rejected = 0, 0
    else:
        # Any comment from a configured IM/Trainer counts as "reviewed".
        # Default guess: acceptance-style language (okay, accepted,
        # approved, ...) -> accepted; any other comment -> rejected. Edit in
        # the review table if the reviewer's actual decision differed.
        combined_comment_text = " ".join(comment for _, comment in official_comments)
        is_accepted = _is_acceptance_comment(combined_comment_text)
        accepted = 1 if is_accepted else 0
        rejected = 0 if is_accepted else 1

    official_names_seen = []
    for name, _ in official_comments:
        canon = name.strip()
        if canon not in official_names_seen:
            official_names_seen.append(canon)

    return {
        "Reviewer(s)": ", ".join(reviewer_names) if reviewer_names else "",
        "Official Reviewers": ", ".join(official_names_seen) if official_names_seen else "",
        "Reviewed by IM": int(reviewed_by_im),
        "Reviewed by Trainer": int(reviewed_by_trainer),
        "Accepted": accepted,
        "Rejected": rejected,
    }


def build_review_table(raw_df: pd.DataFrame, config: ProgramConfig) -> pd.DataFrame:
    """Build the editable intermediate table (mirrors what April's Sheet1
    was doing by hand) from a freshly loaded raw DataFrame."""
    rows = []
    for _, r in raw_df.iterrows():
        classification = classify_row(r.get("Answer"), r.get("Feedback"), config)
        rows.append(
            {
                "NGO Name": r.get("NGO Name"),
                "Category": r.get("Category"),
                "Question": r.get("Question"),
                "Answer": r.get("Answer"),
                "File Name": r.get("File Name"),
                "Feedback": r.get("Feedback"),
                **classification,
            }
        )
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 4. Aggregation: review table -> NGO x Module summary
# --------------------------------------------------------------------------
def aggregate_workflow(review_df: pd.DataFrame, config: ProgramConfig):
    """Aggregate the (possibly user-edited) review table into the NGO x
    Module structure used by the workflow sheet.

    Returns:
        ngo_list: list[str]
        categories: list[str]  (module/category names, in output order)
        params_per_category: dict[str, int]
        table: dict[(ngo, category)] -> dict with uploaded/reviewed_im/
               reviewed_trainer/accepted/rejected
    """
    categories = config.category_order or list(
        dict.fromkeys(review_df["Category"].dropna().tolist())
    )

    params_per_category = dict(config.parameters_per_category)
    for cat in categories:
        if cat not in params_per_category:
            params_per_category[cat] = int(
                review_df.loc[review_df["Category"] == cat, "Question"].nunique()
            )

    ngo_list = sorted(review_df["NGO Name"].dropna().unique().tolist())

    table: dict[tuple[str, str], dict[str, int]] = {}
    for ngo in ngo_list:
        for cat in categories:
            subset = review_df[(review_df["NGO Name"] == ngo) & (review_df["Category"] == cat)]
            table[(ngo, cat)] = {
                "uploaded": int(subset["File Name"].notna().sum()),
                "reviewed_im": int(subset["Reviewed by IM"].sum()),
                "reviewed_trainer": int(subset["Reviewed by Trainer"].sum()),
                "accepted": int(subset["Accepted"].sum()),
                "rejected": int(subset["Rejected"].sum()),
            }

    return ngo_list, categories, params_per_category, table


def reviewer_summary(review_df: pd.DataFrame, config: ProgramConfig):
    """Count how many docs (rows) each configured IM/Trainer actually
    reviewed — i.e. left a comment on, counted once per row even if they
    commented multiple times on the same row.

    Returns a list of dicts: [{"Name": ..., "Role": "IM"/"Trainer",
    "Docs Reviewed": int}, ...], one row per name in config.im_names /
    config.trainer_names, in that order. Names with 0 reviews are still
    included, so you can spot an IM/Trainer who hasn't reviewed anything yet.
    """
    counts = {name.strip(): 0 for name in config.im_names + config.trainer_names}
    im_lookup = {n.strip().lower(): n.strip() for n in config.im_names}
    trainer_lookup = {n.strip().lower(): n.strip() for n in config.trainer_names}

    for reviewers_str in review_df.get("Official Reviewers", pd.Series(dtype=str)).dropna():
        if not reviewers_str:
            continue
        for name in reviewers_str.split(", "):
            name = name.strip()
            if not name:
                continue
            key = name.lower()
            canon = im_lookup.get(key) or trainer_lookup.get(key)
            if canon:
                counts[canon] = counts.get(canon, 0) + 1

    rows = []
    for name in config.im_names:
        rows.append({"Name": name.strip(), "Role": "IM", "Docs Reviewed": counts.get(name.strip(), 0)})
    for name in config.trainer_names:
        rows.append({"Name": name.strip(), "Role": "Trainer", "Docs Reviewed": counts.get(name.strip(), 0)})
    return rows


# --------------------------------------------------------------------------
# 5. Excel writer — replicates the workflow-sheet layout
# --------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_workflow_excel(
    ngo_list,
    categories,
    params_per_category,
    table,
    config: ProgramConfig,
    reviewer_summary_rows=None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = config.sheet_name

    sub_cols = [
        "# of evidence docs uploaded",
        "# of evidence docs reviewed and commented on by IMs",
        "# of evidence docs reviewed and commented on by Trainers",
        "# accepted",
        "# rejected",
    ]
    n_sub = len(sub_cols)

    first_data_col = 2  # column B, to mirror the original sheet's layout
    ngo_col = first_data_col

    # Row 1: assessment name
    ws.cell(row=1, column=ngo_col, value=config.assessment_name)
    # Row 2: report date
    ws.cell(row=2, column=ngo_col, value=f"Report Date: {config.report_date}")

    # Row 4: Module Name header (merged across each category's sub-columns)
    module_row = 4
    params_row = 5
    subheader_row = 6
    first_ngo_row = 7

    ws.cell(row=module_row, column=ngo_col, value="Module Name").font = BOLD
    ws.cell(row=params_row, column=ngo_col, value="# of parameters").font = BOLD
    ws.cell(row=subheader_row, column=ngo_col, value="NGO").font = BOLD

    col = ngo_col + 1
    total_start_col = None
    for cat in categories:
        ws.cell(row=module_row, column=col, value=cat)
        ws.merge_cells(start_row=module_row, start_column=col, end_row=module_row, end_column=col + n_sub - 1)
        ws.cell(row=module_row, column=col).alignment = CENTER
        ws.cell(row=module_row, column=col).font = BOLD
        ws.cell(row=module_row, column=col).fill = HEADER_FILL

        ws.cell(row=params_row, column=col, value=params_per_category.get(cat, 0))
        ws.merge_cells(start_row=params_row, start_column=col, end_row=params_row, end_column=col + n_sub - 1)
        ws.cell(row=params_row, column=col).alignment = CENTER

        for j, sub in enumerate(sub_cols):
            c = ws.cell(row=subheader_row, column=col + j, value=sub)
            c.font = BOLD
            c.alignment = CENTER
            c.fill = HEADER_FILL
        col += n_sub

    total_start_col = col
    total_headers = [
        "Total # of evidence docs uploaded",
        "% of Docs Uploaded",
        "Total # of evidence docs reviewed and commented on by IMs",
        "% of Docs Reviewed by IMs",
        "Total # of evidence docs reviewed and commented on by Trainers",
        "% of Docs Reviewed by Trainers",
        "Total # accepted",
        "Total # rejected",
    ]
    ws.cell(row=module_row, column=total_start_col, value="Total Number of Documents")
    ws.merge_cells(
        start_row=module_row,
        start_column=total_start_col,
        end_row=module_row,
        end_column=total_start_col + len(total_headers) - 1,
    )
    ws.cell(row=module_row, column=total_start_col).font = BOLD
    ws.cell(row=module_row, column=total_start_col).alignment = CENTER
    ws.cell(row=module_row, column=total_start_col).fill = HEADER_FILL

    for j, h in enumerate(total_headers):
        c = ws.cell(row=subheader_row, column=total_start_col + j, value=h)
        c.font = BOLD
        c.alignment = CENTER
        c.fill = HEADER_FILL

    total_params = sum(params_per_category.get(cat, 0) for cat in categories)

    # Data rows
    grand_totals = {h: 0 for h in ["uploaded", "reviewed_im", "reviewed_trainer", "accepted", "rejected"]}
    row_idx = first_ngo_row
    for ngo in ngo_list:
        ws.cell(row=row_idx, column=ngo_col, value=ngo)
        col = ngo_col + 1
        row_total = {"uploaded": 0, "reviewed_im": 0, "reviewed_trainer": 0, "accepted": 0, "rejected": 0}
        for cat in categories:
            vals = table.get((ngo, cat), {"uploaded": 0, "reviewed_im": 0, "reviewed_trainer": 0, "accepted": 0, "rejected": 0})
            ordered = [vals["uploaded"], vals["reviewed_im"], vals["reviewed_trainer"], vals["accepted"], vals["rejected"]]
            for j, v in enumerate(ordered):
                ws.cell(row=row_idx, column=col + j, value=v)
            for k in row_total:
                row_total[k] += vals[k]
            col += n_sub

        pct_uploaded = round(row_total["uploaded"] / total_params, 4) if total_params else 0
        pct_reviewed_im = round(row_total["reviewed_im"] / row_total["uploaded"], 4) if row_total["uploaded"] else 0
        pct_reviewed_trainer = round(row_total["reviewed_trainer"] / row_total["uploaded"], 4) if row_total["uploaded"] else 0

        totals_row_vals = [
            row_total["uploaded"],
            pct_uploaded,
            row_total["reviewed_im"],
            pct_reviewed_im,
            row_total["reviewed_trainer"],
            pct_reviewed_trainer,
            row_total["accepted"],
            row_total["rejected"],
        ]
        for j, v in enumerate(totals_row_vals):
            ws.cell(row=row_idx, column=total_start_col + j, value=v)

        for k in grand_totals:
            grand_totals[k] += row_total[k]

        row_idx += 1

    # Grand total row
    ws.cell(row=row_idx, column=ngo_col, value="TOTAL").font = BOLD
    col = ngo_col + 1
    for cat in categories:
        # per-category grand totals (sum down the column)
        cat_totals = {"uploaded": 0, "reviewed_im": 0, "reviewed_trainer": 0, "accepted": 0, "rejected": 0}
        for ngo in ngo_list:
            v = table.get((ngo, cat))
            if v:
                for k in cat_totals:
                    cat_totals[k] += v[k]
        ordered = [cat_totals["uploaded"], cat_totals["reviewed_im"], cat_totals["reviewed_trainer"], cat_totals["accepted"], cat_totals["rejected"]]
        for j, v in enumerate(ordered):
            cell = ws.cell(row=row_idx, column=col + j, value=v)
            cell.font = BOLD
            cell.fill = TOTAL_FILL
        col += n_sub

    pct_uploaded_total = round(grand_totals["uploaded"] / total_params, 4) if total_params else 0
    pct_reviewed_im_total = round(grand_totals["reviewed_im"] / grand_totals["uploaded"], 4) if grand_totals["uploaded"] else 0
    pct_reviewed_trainer_total = round(grand_totals["reviewed_trainer"] / grand_totals["uploaded"], 4) if grand_totals["uploaded"] else 0
    grand_vals = [
        grand_totals["uploaded"],
        pct_uploaded_total,
        grand_totals["reviewed_im"],
        pct_reviewed_im_total,
        grand_totals["reviewed_trainer"],
        pct_reviewed_trainer_total,
        grand_totals["accepted"],
        grand_totals["rejected"],
    ]
    for j, v in enumerate(grand_vals):
        cell = ws.cell(row=row_idx, column=total_start_col + j, value=v)
        cell.font = BOLD
        cell.fill = TOTAL_FILL

    # Column widths
    ws.column_dimensions[get_column_letter(ngo_col)].width = 38
    for c in range(ngo_col + 1, total_start_col + len(total_headers)):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # --- Reviewer Summary sheet: # of docs reviewed per individual IM/Trainer ---
    if reviewer_summary_rows:
        ws2 = wb.create_sheet("Reviewer Summary")
        headers = ["Name", "Role", "Docs Reviewed"]
        for j, h in enumerate(headers, start=1):
            c = ws2.cell(row=1, column=j, value=h)
            c.font = BOLD
            c.fill = HEADER_FILL
            c.alignment = CENTER
        for i, row in enumerate(reviewer_summary_rows, start=2):
            ws2.cell(row=i, column=1, value=row["Name"])
            ws2.cell(row=i, column=2, value=row["Role"])
            ws2.cell(row=i, column=3, value=row["Docs Reviewed"])
        ws2.column_dimensions["A"].width = 24
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 16

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
